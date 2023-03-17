import openpyxl

#
class LoginPageData:

    @staticmethod
    def getTestData(test_case_name):
        Dict={}
        workbook = openpyxl.load_workbook("C:/Users/DELL/Desktop/Assignment/Testdata.xlsx")
        sheet = workbook.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2, sheet.max_column + 1):  # to get columns
                    # Dict["lastname"]="shetty
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        print(Dict)
        return [Dict]


# Dict={}
# workbook = openpyxl.load_workbook("C:/Users/DELL/Desktop/Assignment/Testdata.xlsx")
# sheet = workbook.active
# for i in range(1, sheet.max_row):  # to get rows
#     if sheet.cell(row=i, column=1).value:
#         for j in range(2, sheet.max_column):
#             Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
# print(Dict)










